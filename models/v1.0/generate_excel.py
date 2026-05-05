import json
import pandas as pd
from collections import defaultdict
from openpyxl.utils import get_column_letter

# Загрузка данных
with open('genereted/solver_results.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

requests = data['requests']
edges = data['edges']

# ============================================================================
# Файл 1: Матрица S×C (источники × потребители)
# ============================================================================

sources = sorted(set(r['source'] for r in requests))
consumers = sorted(set(r['consumer'] for r in requests), key=lambda x: int(x) if x.isdigit() else x)

matrix_s_c = defaultdict(dict)
for r in requests:
    src = r['source']
    cons = r['consumer']
    delivered = r['delivered']
    demanded = r['demanded']
    pct = r['delivery_pct']
    matrix_s_c[src][cons] = f"{delivered:.1f} / {demanded:.1f}\n{pct:.1f}%"

df1 = pd.DataFrame(index=sources, columns=consumers)
for src in sources:
    for cons in consumers:
        df1.loc[src, cons] = matrix_s_c.get(src, {}).get(cons, "-")

df1.index.name = "Источник"
df1.columns.name = "Потребитель"

with pd.ExcelWriter('genereted/matrix_sources_consumers.xlsx') as writer:
    df1.to_excel(writer, sheet_name='Доставка', merge_cells=False)
    ws = writer.sheets['Доставка']
    ws.column_dimensions['A'].width = 12
    for i, col in enumerate(consumers, 1):
        ws.column_dimensions[get_column_letter(i + 1)].width = 18

print("✓ Файл 1: genereted/matrix_sources_consumers.xlsx")

# ============================================================================
# Файл 2: Матрица E×E (рёбра)
# ============================================================================

all_nodes = set()
edge_dict = {}

for e in edges:
    parts = e['edge'].split(' ↔ ')
    n1, n2 = parts[0], parts[1]
    all_nodes.add(n1)
    all_nodes.add(n2)
    flow = e['flow']
    cap = e['capacity']
    util = e['utilization']
    edge_dict[(n1, n2)] = (flow, cap, util)
    edge_dict[(n2, n1)] = (flow, cap, util)

nodes = sorted(all_nodes, key=lambda x: (not x.isdigit(), not x.isalpha(), x))

df2 = pd.DataFrame(index=nodes, columns=nodes)

for (n1, n2), (flow, cap, util) in edge_dict.items():
    if cap == 'inf':
        df2.loc[n1, n2] = f"{flow:.1f} / inf\n{util:.1f}%"
    else:
        cap_float = float(cap) if isinstance(cap, str) else cap
        df2.loc[n1, n2] = f"{flow:.1f} / {cap_float:.1f}\n{util:.1f}%"

df2 = df2.fillna("—")

with pd.ExcelWriter('genereted/matrix_edges.xlsx') as writer:
    df2.to_excel(writer, sheet_name='Рёбра', merge_cells=False)
    ws = writer.sheets['Рёбра']
    ws.column_dimensions['A'].width = 12
    for i in range(len(nodes)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 18

print("✓ Файл 2: genereted/matrix_edges.xlsx")